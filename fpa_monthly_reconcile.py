from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Literal, Sequence

import pdfplumber
from openpyxl import Workbook, load_workbook

try:
    from rapidfuzz import fuzz
except ImportError:  # The tool works without extra installs.
    fuzz = None

from difflib import SequenceMatcher


MatchMode = Literal["safe", "review", "code_only", "code_then_name"]
DATE_RE = re.compile(r"^(\d{1,2}-[A-Za-z]{3}-\d{2})\s+")
TXN_RE = re.compile(
    r"^(?:(?P<date>\d{1,2}-[A-Za-z]{3}-\d{2})\s+)?"
    r"By\s+\(as\s+per\s+details\)\s+B2C\s+Sales\s+"
    r"(?P<voucher>\S+)\s+(?P<amount>[\d,]+(?:\.\d{1,2})?)"
)
CODE_RE = re.compile(r"\bIBOC\s*\d+(?:\s*-\s*\d+){0,2}\b", re.IGNORECASE)
MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


@dataclass
class PdfTransaction:
    index: int
    source_pdf: str
    date_text: str
    date_value: datetime | None
    voucher: str
    student_code: str
    student_name: str
    received_amount: float | None
    net_amount: float | None
    page: int
    raw_text: str


@dataclass
class ExcelRow:
    row_number: int
    student_code: str
    student_name: str
    fees_received_with_gst: float | None


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def normalize_code(value: object) -> str:
    text = str(value or "").upper().strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"-+", "-", text)
    return text


def normalize_name(value: object) -> str:
    text = str(value or "").lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_amount(value: str | None) -> float | None:
    if not value:
        return None
    try:
        return round(float(value.replace(",", "")), 2)
    except ValueError:
        return None


def parse_excel_amount(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip()
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def amounts_match(left: float | None, right: float | None, tolerance: float = 1.0) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def parse_tally_date(value: str) -> datetime | None:
    match = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{2})$", value.strip())
    if not match:
        return None
    day = int(match.group(1))
    month = MONTHS.get(match.group(2).lower())
    year = 2000 + int(match.group(3))
    if not month:
        return None
    return datetime(year, month, day)


def similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if fuzz:
        return float(fuzz.token_sort_ratio(left, right))
    return SequenceMatcher(None, " ".join(sorted(left.split())), " ".join(sorted(right.split()))).ratio() * 100


def clean_student_code(block_text: str) -> str:
    matches = CODE_RE.findall(block_text)
    if not matches:
        return ""
    return normalize_code(matches[-1])


def clean_student_name(block_text: str, code: str) -> str:
    text = re.sub(r"\s+", " ", block_text).strip()
    if code:
        flexible_code = r"\s*-\s*".join(map(re.escape, code.split("-")))
        text = re.sub(flexible_code, "", text, flags=re.IGNORECASE)

    match = re.search(r"(.+?)\s*-\s*IBOC\s+fee+s?\b", text, flags=re.IGNORECASE)
    if match:
        name = match.group(1)
    else:
        code_match = CODE_RE.search(text)
        name = text[: code_match.start()] if code_match else text
        name = name.split(" - ")[0]

    name = re.sub(r"\bBy\b|\bHDFC\b|\bCGST\b|\bSGST\b|PAYABLE|Cr|Dr", " ", name, flags=re.IGNORECASE)
    name = re.sub(r"\d[\d,]*(?:\.\d+)?", " ", name)
    name = re.sub(r"\s+", " ", name).strip(" -")
    return name


def iter_pdf_blocks(pdf_path: Path) -> Iterable[tuple[int, str, str, str, float | None, float | None, str]]:
    last_seen_date = ""
    current_txn_date = ""
    current_page = 0
    current_voucher = ""
    current_net_amount: float | None = None
    current_received_amount: float | None = None
    student_lines: list[str] = []
    raw_lines: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                date_match = DATE_RE.match(line)
                if date_match:
                    last_seen_date = date_match.group(1)

                txn_match = TXN_RE.match(line)
                if txn_match:
                    if raw_lines and current_voucher:
                        yield (
                            current_page,
                            current_txn_date,
                            current_voucher,
                            " ".join(student_lines),
                            current_net_amount,
                            current_received_amount,
                            "\n".join(raw_lines),
                        )
                    current_page = page_number
                    current_txn_date = txn_match.group("date") or last_seen_date
                    last_seen_date = current_txn_date
                    current_voucher = txn_match.group("voucher")
                    current_net_amount = parse_amount(txn_match.group("amount"))
                    current_received_amount = None
                    student_lines = []
                    raw_lines = [line]
                    continue

                if current_voucher:
                    raw_lines.append(line)
                    received_match = re.search(r"\b(?:HDFC|BANK|ICICI|AXIS|SBI)\b.*?([\d,]+(?:\.\d{1,2})?)\s+Dr\b", line, re.IGNORECASE)
                    if received_match:
                        current_received_amount = parse_amount(received_match.group(1))

                    skip = (
                        line.startswith("HDFC ")
                        or line.upper().startswith("BANK ")
                        or line.upper().startswith("ICICI ")
                        or line.upper().startswith("AXIS ")
                        or line.upper().startswith("SBI ")
                        or line.startswith("CGST ")
                        or line.startswith("SGST ")
                        or line.startswith("Carried Over")
                        or line.startswith("Brought Forward")
                        or line.startswith("continued")
                        or line.startswith("Date Particulars")
                    )
                    if not skip:
                        student_lines.append(line)

    if raw_lines and current_voucher:
        yield (
            current_page,
            current_txn_date,
            current_voucher,
            " ".join(student_lines),
            current_net_amount,
            current_received_amount,
            "\n".join(raw_lines),
        )


def extract_pdf_transactions(pdf_path: Path) -> list[PdfTransaction]:
    transactions: list[PdfTransaction] = []
    for page, date_text, voucher, block_text, net_amount, received_amount, raw_text in iter_pdf_blocks(pdf_path):
        code = clean_student_code(block_text)
        name = clean_student_name(block_text, code)
        if not code and not name:
            continue
        transactions.append(
            PdfTransaction(
                index=len(transactions) + 1,
                source_pdf=pdf_path.name,
                date_text=date_text,
                date_value=parse_tally_date(date_text),
                voucher=voucher,
                student_code=code,
                student_name=name,
                received_amount=received_amount,
                net_amount=net_amount,
                page=page,
                raw_text=raw_text,
            )
        )
    return transactions


def normalize_pdf_paths(pdf_paths: Path | Sequence[Path]) -> list[Path]:
    if isinstance(pdf_paths, Path):
        return [pdf_paths]
    return list(pdf_paths)


def extract_pdf_transactions_from_paths(pdf_paths: Path | Sequence[Path]) -> list[PdfTransaction]:
    transactions: list[PdfTransaction] = []
    for pdf_path in normalize_pdf_paths(pdf_paths):
        for txn in extract_pdf_transactions(pdf_path):
            txn.index = len(transactions) + 1
            transactions.append(txn)
    return transactions


def find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        "code": {"studentcode", "code"},
        "name": {"studentname", "name", "namevendor", "namevendore"},
        "fees_received": {"feesreceivedwithgst", "feesreceived", "amountreceivedwithgst", "receivedamountwithgst"},
        "tally_date": {"tallydate"},
        "tally_voucher": {"tallyvoucher", "tallyvchno", "tallyvoucherno"},
    }
    for row in range(1, min(ws.max_row, 25) + 1):
        normalized = {normalize_header(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1)}
        found: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                if name in normalized:
                    found[key] = normalized[name]
                    break
        if {"code", "name", "fees_received", "tally_date", "tally_voucher"}.issubset(found):
            return row, found
    raise ValueError(
        "Could not find required Excel headers. Expected Code/Student Code, "
        "Name / Vendor/Student Name, Fees Received with GST, Tally Date, and Tally Voucher."
    )


def read_excel_rows(ws, header_row: int, columns: dict[str, int]) -> list[ExcelRow]:
    rows: list[ExcelRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        code = normalize_code(ws.cell(row, columns["code"]).value)
        name = str(ws.cell(row, columns["name"]).value or "").strip()
        fees_received = parse_excel_amount(ws.cell(row, columns["fees_received"]).value)
        if code or name:
            rows.append(ExcelRow(row, code, name, fees_received))
    return rows


def canonical_match_mode(match_mode: MatchMode) -> Literal["safe", "review"]:
    if match_mode in {"review", "code_then_name"}:
        return "review"
    return "safe"


def choose_match(row: ExcelRow, transactions: list[PdfTransaction], used_txns: set[int], match_mode: MatchMode):
    mode = canonical_match_mode(match_mode)
    available = [txn for txn in transactions if txn.index not in used_txns]
    code_matches = [txn for txn in available if row.student_code and txn.student_code == row.student_code]
    code_amount_matches = [
        txn for txn in code_matches if amounts_match(row.fees_received_with_gst, txn.received_amount)
    ]
    if len(code_amount_matches) == 1:
        return "matched", code_amount_matches[0], "Priority 1", 100.0, [], "Student code and amount matched"
    if len(code_amount_matches) > 1:
        return (
            "duplicate",
            None,
            "Priority 1",
            100.0,
            code_amount_matches,
            "Multiple PDF vouchers matched the same student code and amount",
        )

    if len(code_matches) == 1:
        return "matched", code_matches[0], "Priority 2", 96.0, [], "Unique student code matched"

    if mode == "safe":
        if len(code_matches) > 1:
            return (
                "duplicate",
                None,
                "Priority 2",
                90.0,
                code_matches,
                "Multiple PDF vouchers matched the same student code; amount did not identify one voucher",
            )
        return "unmatched", None, "No safe code or code+amount match found", 0.0, [], "No safe code or code+amount match found"

    row_name_norm = normalize_name(row.student_name)
    code_name_matches = [
        txn for txn in code_matches if row_name_norm and normalize_name(txn.student_name) == row_name_norm
    ]
    if len(code_name_matches) == 1:
        return "matched", code_name_matches[0], "Priority 3", 95.0, [], "Student code and exact name matched"
    if len(code_name_matches) > 1:
        return (
            "duplicate",
            None,
            "Priority 3",
            90.0,
            code_name_matches,
            "Multiple PDF vouchers matched the same student code and exact name",
        )

    exact_name_amount_matches = [
        txn
        for txn in available
        if row_name_norm
        and normalize_name(txn.student_name) == row_name_norm
        and amounts_match(row.fees_received_with_gst, txn.received_amount)
    ]
    if len(exact_name_amount_matches) == 1:
        return "matched", exact_name_amount_matches[0], "Priority 4", 94.0, [], "Exact name and amount matched"
    if len(exact_name_amount_matches) > 1:
        return (
            "duplicate",
            None,
            "Priority 4",
            90.0,
            exact_name_amount_matches,
            "Multiple PDF vouchers matched the same exact name and amount",
        )

    fuzzy_matches = []
    for txn in available:
        score = similarity(row_name_norm, normalize_name(txn.student_name))
        if score > 95 and amounts_match(row.fees_received_with_gst, txn.received_amount):
            fuzzy_matches.append((score, txn))
    fuzzy_matches.sort(reverse=True, key=lambda item: item[0])
    if len(fuzzy_matches) == 1:
        score, txn = fuzzy_matches[0]
        return "matched", txn, "Priority 5", score, [], "Fuzzy name above 95% and amount matched"
    if len(fuzzy_matches) > 1:
        return (
            "duplicate",
            None,
            "Priority 5",
            fuzzy_matches[0][0],
            [x[1] for x in fuzzy_matches],
            "Multiple PDF vouchers matched by fuzzy name and amount",
        )

    best_score = 0.0
    for txn in available:
        best_score = max(best_score, similarity(row_name_norm, normalize_name(txn.student_name)))
    if len(code_matches) > 1:
        return (
            "duplicate",
            None,
            "Manual Review",
            90.0,
            code_matches,
            "Multiple PDF vouchers matched the same student code; review mode could not safely choose one",
        )
    return (
        "unmatched",
        None,
        "No safe match found",
        best_score,
        [],
        "No code+amount, unique code, exact name+amount, or fuzzy name+amount match found",
    )


def append_report_rows(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)


def save_report(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    append_report_rows(ws, headers, rows)
    wb.save(path)


def export_pdf_tally_details(pdf_paths: Path | Sequence[Path], output_dir: Path) -> dict[str, int | Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path_list = normalize_pdf_paths(pdf_paths)
    transactions = extract_pdf_transactions_from_paths(pdf_path_list)
    rows = [
        [
            txn.date_value or txn.date_text,
            txn.voucher,
            txn.source_pdf,
            txn.student_code,
            txn.student_name,
            txn.received_amount,
            txn.net_amount,
            txn.page,
        ]
        for txn in transactions
    ]

    output_stem = pdf_path_list[0].stem if len(pdf_path_list) == 1 else "combined"
    output_path = output_dir / f"{output_stem}_tally_date_voucher.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Date Voucher"
    append_report_rows(
        ws,
        [
            "Tally Date",
            "Tally Voucher",
            "Source PDF",
            "Student Code",
            "Student Name",
            "Received Amount with GST",
            "Net Amount",
            "PDF Page",
        ],
        rows,
    )
    ws.freeze_panes = "A2"
    wb.save(output_path)
    return {"transactions": len(transactions), "pdf_files": len(pdf_path_list), "output_path": output_path}


def reconcile(
    pdf_path: Path | Sequence[Path],
    excel_path: Path,
    output_dir: Path,
    match_mode: MatchMode = "code_only",
) -> dict[str, int | Path | str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path_list = normalize_pdf_paths(pdf_path)
    transactions = extract_pdf_transactions_from_paths(pdf_path_list)

    wb = load_workbook(excel_path)
    ws = wb.active
    header_row, columns = find_header_row_and_columns(ws)
    excel_rows = read_excel_rows(ws, header_row, columns)

    used_txns: set[int] = set()
    manual_review_txns: set[int] = set()
    decided_rows: set[int] = set()
    matched_rows: list[list[object]] = []
    unmatched_rows: list[list[object]] = []
    duplicate_rows: list[list[object]] = []

    def available_transactions() -> list[PdfTransaction]:
        return [txn for txn in transactions if txn.index not in used_txns]

    def write_match(row: ExcelRow, txn: PdfTransaction, priority: str, confidence: float, reason: str) -> None:
        ws.cell(row.row_number, columns["tally_date"]).value = txn.date_value or txn.date_text
        ws.cell(row.row_number, columns["tally_voucher"]).value = txn.voucher
        used_txns.add(txn.index)
        decided_rows.add(row.row_number)
        matched_rows.append(
            [
                row.row_number,
                row.student_code,
                row.student_name,
                row.fees_received_with_gst,
                txn.source_pdf,
                txn.date_text,
                txn.voucher,
                txn.student_code,
                txn.student_name,
                txn.received_amount,
                priority,
                round(confidence, 2),
                reason,
                txn.page,
                txn.raw_text,
            ]
        )

    def write_duplicate(
        row: ExcelRow,
        possible: list[PdfTransaction],
        reason: str,
    ) -> None:
        if row.row_number in decided_rows:
            return
        decided_rows.add(row.row_number)
        for item in possible:
            manual_review_txns.add(item.index)
        duplicate_rows.append(
            [
                row.student_code,
                row.student_name,
                row.row_number,
                row.fees_received_with_gst,
                "; ".join(
                    f"{item.source_pdf} page {item.page}: {item.date_text} / {item.voucher} / {item.student_code} / {item.student_name} / {item.received_amount}"
                    for item in possible
                ),
                "; ".join(str(item.received_amount) for item in possible),
                reason,
            ]
        )

    def run_unique_candidate_pass(
        priority: str,
        confidence: float,
        reason: str,
        candidate_func,
        duplicate_on_multiple: bool = True,
    ) -> None:
        proposals: dict[int, tuple[ExcelRow, list[PdfTransaction]]] = {}
        for row in excel_rows:
            if row.row_number in decided_rows:
                continue
            candidates = candidate_func(row, available_transactions())
            if candidates:
                proposals[row.row_number] = (row, candidates)

        single_candidates = {
            row_number: (row, candidates[0])
            for row_number, (row, candidates) in proposals.items()
            if len(candidates) == 1
        }
        txn_claims: dict[int, list[ExcelRow]] = {}
        for row, txn in single_candidates.values():
            txn_claims.setdefault(txn.index, []).append(row)

        for row_number, (row, candidates) in proposals.items():
            if row.row_number in decided_rows:
                continue
            if len(candidates) > 1:
                if duplicate_on_multiple:
                    write_duplicate(row, candidates, f"{reason}; multiple possible PDF vouchers")
                continue
            txn = candidates[0]
            claimants = txn_claims.get(txn.index, [])
            if len(claimants) == 1 and txn.index not in used_txns:
                write_match(row, txn, priority, confidence, reason)
            elif len(claimants) > 1:
                write_duplicate(row, [txn], f"{reason}; multiple Excel rows claimed the same PDF voucher")

    run_unique_candidate_pass(
        "Priority 1",
        100.0,
        "Student code and Fees Received with GST amount matched",
        lambda row, txns: [
            txn
            for txn in txns
            if row.student_code
            and txn.student_code == row.student_code
            and amounts_match(row.fees_received_with_gst, txn.received_amount)
        ],
    )

    run_unique_candidate_pass(
        "Priority 2",
        96.0,
        "Unique student code matched after amount matches were processed",
        lambda row, txns: [txn for txn in txns if row.student_code and txn.student_code == row.student_code],
        duplicate_on_multiple=canonical_match_mode(match_mode) == "safe",
    )

    if canonical_match_mode(match_mode) == "review":
        run_unique_candidate_pass(
            "Priority 3",
            95.0,
            "Student code and exact student name matched",
            lambda row, txns: [
                txn
                for txn in txns
                if row.student_code
                and txn.student_code == row.student_code
                and normalize_name(row.student_name)
                and normalize_name(txn.student_name) == normalize_name(row.student_name)
            ],
        )
        run_unique_candidate_pass(
            "Priority 4",
            94.0,
            "Exact student name and Fees Received with GST amount matched",
            lambda row, txns: [
                txn
                for txn in txns
                if normalize_name(row.student_name)
                and normalize_name(txn.student_name) == normalize_name(row.student_name)
                and amounts_match(row.fees_received_with_gst, txn.received_amount)
            ],
        )
        run_unique_candidate_pass(
            "Priority 5",
            95.0,
            "Fuzzy student name above 95% and Fees Received with GST amount matched",
            lambda row, txns: [
                txn
                for txn in txns
                if similarity(normalize_name(row.student_name), normalize_name(txn.student_name)) > 95
                and amounts_match(row.fees_received_with_gst, txn.received_amount)
            ],
        )

    for row in excel_rows:
        if row.row_number in decided_rows:
            continue
        best_score = max(
            [similarity(normalize_name(row.student_name), normalize_name(txn.student_name)) for txn in transactions],
            default=0.0,
        )
        unmatched_rows.append(
            [
                row.row_number,
                row.student_code,
                row.student_name,
                row.fees_received_with_gst,
                "No safe match found after all enabled matching passes",
                round(best_score, 2),
            ]
        )

    pdf_unmatched_rows: list[list[object]] = []
    for txn in transactions:
        if txn.index in used_txns or txn.index in manual_review_txns:
            continue
        pdf_unmatched_rows.append(
            [
                txn.source_pdf,
                txn.date_text,
                txn.voucher,
                txn.student_code,
                txn.student_name,
                txn.received_amount,
                txn.net_amount,
                txn.page,
                "PDF voucher was not used by any safe Excel match and was not part of manual review",
                txn.raw_text,
            ]
        )

    stem = excel_path.stem
    updated_path = output_dir / f"{stem}_updated.xlsx"
    matched_path = output_dir / "matched_report.xlsx"
    excel_unmatched_path = output_dir / "excel_unmatched_report.xlsx"
    pdf_unmatched_path = output_dir / "pdf_unmatched_report.xlsx"
    duplicate_path = output_dir / "duplicate_manual_review_report.xlsx"
    summary_path = output_dir / "summary_report.xlsx"

    wb.save(updated_path)
    save_report(
        matched_path,
        [
            "Excel Row",
            "Student Code",
            "Student Name",
            "Excel Fees Received with GST",
            "Source PDF",
            "PDF Date",
            "PDF Voucher Number",
            "PDF Student Code",
            "PDF Student Name",
            "PDF Received Amount",
            "Match Priority",
            "Confidence",
            "Match Reason",
            "PDF Page",
            "Raw PDF Transaction Text",
        ],
        matched_rows,
    )
    save_report(
        excel_unmatched_path,
        ["Excel Row", "Student Code", "Student Name", "Fees Received with GST", "Reason", "Best Name Confidence"],
        unmatched_rows,
    )
    save_report(
        pdf_unmatched_path,
        [
            "Source PDF",
            "PDF Date",
            "PDF Voucher Number",
            "PDF Student Code",
            "PDF Student Name",
            "PDF Received Amount",
            "PDF Net Amount",
            "PDF Page Number",
            "Reason",
            "Raw PDF Transaction Text",
        ],
        pdf_unmatched_rows,
    )
    save_report(
        duplicate_path,
        [
            "Student Code",
            "Student Name",
            "Excel Row(s)",
            "Excel Fees Received with GST",
            "PDF Voucher(s)",
            "Amount(s)",
            "Reason manual review is needed",
        ],
        duplicate_rows,
    )
    pdf_manual_review_count = len(manual_review_txns)
    pdf_matched_count = len(used_txns)
    summary_rows = [
        ["Match mode", "Safe mode" if canonical_match_mode(match_mode) == "safe" else "Review mode"],
        ["PDF files uploaded", len(pdf_path_list)],
        ["PDF transactions extracted", len(transactions)],
        ["PDF transactions matched", pdf_matched_count],
        ["PDF transactions unmatched", len(pdf_unmatched_rows)],
        ["PDF transactions needing manual review", pdf_manual_review_count],
        ["Excel rows scanned", len(excel_rows)],
        ["Excel rows updated", len(matched_rows)],
        ["Excel rows unmatched", len(unmatched_rows)],
        ["Duplicate/manual review count", len(duplicate_rows)],
        ["Updated workbook", str(updated_path)],
        ["Matched report", str(matched_path)],
        ["Excel unmatched report", str(excel_unmatched_path)],
        ["PDF unmatched report", str(pdf_unmatched_path)],
        ["Duplicate/manual review report", str(duplicate_path)],
    ]
    save_report(summary_path, ["Metric", "Value"], summary_rows)

    return {
        "pdf_transactions": len(transactions),
        "pdf_files": len(pdf_path_list),
        "excel_rows": len(excel_rows),
        "matched": len(matched_rows),
        "unmatched": len(unmatched_rows),
        "duplicates": len(duplicate_rows),
        "pdf_matched": pdf_matched_count,
        "pdf_unmatched": len(pdf_unmatched_rows),
        "pdf_manual_review": pdf_manual_review_count,
        "match_mode": match_mode,
        "updated_path": updated_path,
        "matched_path": matched_path,
        "excel_unmatched_path": excel_unmatched_path,
        "pdf_unmatched_path": pdf_unmatched_path,
        "duplicate_path": duplicate_path,
        "summary_path": summary_path,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill Tally Date and Tally Voucher from one or more Tally PDFs.")
    parser.add_argument("--pdf", required=True, nargs="+", type=Path, help="Path(s) to Tally ledger PDF files")
    parser.add_argument("--excel", required=True, type=Path, help="Path to the student accounting workbook")
    parser.add_argument("--output-dir", default=Path("outputs"), type=Path, help="Folder for generated files")
    parser.add_argument(
        "--match-mode",
        choices=["safe", "review", "code_only", "code_then_name"],
        default="safe",
        help="Use safe for code+amount then unique code, or review to also allow name+amount matching.",
    )
    args = parser.parse_args()

    result = reconcile(args.pdf, args.excel, args.output_dir, args.match_mode)
    print("FPA monthly reconciliation complete")
    print(f"PDF transactions extracted: {result['pdf_transactions']}")
    print(f"PDF transactions matched: {result['pdf_matched']}")
    print(f"PDF transactions unmatched: {result['pdf_unmatched']}")
    print(f"PDF transactions needing manual review: {result['pdf_manual_review']}")
    print(f"Excel rows scanned: {result['excel_rows']}")
    print(f"Excel rows updated: {result['matched']}")
    print(f"Excel rows unmatched: {result['unmatched']}")
    print(f"Duplicate/manual review rows: {result['duplicates']}")
    print(f"Updated workbook: {result['updated_path']}")
    print(f"Reports folder: {args.output_dir}")


if __name__ == "__main__":
    main()
