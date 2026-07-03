from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


TALLY_DATE_COL = 37  # AK
TALLY_VOUCHER_COL = 38  # AL
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


@dataclass
class PostingOptions:
    target_sheet: str = "IS25-26"
    allow_overwrite: bool = False
    highlight_review_rows: bool = True
    require_verified_status: bool = True
    require_amount_match: bool = True


@dataclass
class ReconRecord:
    row_number: int
    student_code: str
    student_name: str
    course: str
    amount: float | None
    tally_date: Any
    tally_voucher: Any
    match_status: str
    verified_status: str
    final_row: int | None


@dataclass
class IncomeRecord:
    row_number: int
    student_code: str
    student_name: str
    course: str
    amount: float | None


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def normalize_text(value: object) -> str:
    text = str(value or "").lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_code(value: object) -> str:
    text = str(value or "").upper().strip()
    text = re.sub(r"\s+", "", text)
    return re.sub(r"-+", "-", text)


def parse_amount(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = re.sub(r"[^\d.\-]", "", str(value))
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def amounts_match(left: float | None, right: float | None, tolerance: float = 1.0) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def is_approved_status(value: str) -> bool:
    if not value:
        return True
    normalized = normalize_text(value)
    approved = {"matched", "safe", "approved", "yes", "verified", "posted", "ok"}
    review_words = {"unmatched", "duplicate", "review", "pending", "no", "failed", "mismatch"}
    if any(word in normalized.split() for word in review_words):
        return False
    return any(word in normalized.split() for word in approved)


def find_header_row(ws, required_aliases: dict[str, set[str]], max_scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_found: dict[str, int] = {}
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        headers = {normalize_header(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1)}
        found: dict[str, int] = {}
        for key, aliases in required_aliases.items():
            for alias in aliases:
                if alias in headers:
                    found[key] = headers[alias]
                    break
        if len(found) > len(best_found):
            best_row = row
            best_found = found
        if "tally_date" in found and "tally_voucher" in found and ("student_code" in found or "student_name" in found):
            return row, found
    return best_row, best_found


def recon_aliases() -> dict[str, set[str]]:
    return {
        "student_code": {"studentcode", "code", "pdfstudentcode"},
        "student_name": {"studentname", "name", "namevendor", "pdfstudentname"},
        "course": {"course", "coursename", "program"},
        "amount": {"amount", "feesreceivedwithgst", "excelfeesreceivedwithgst", "pdfreceivedamount"},
        "tally_date": {"tallydate", "pdfdate", "tallydatefound"},
        "tally_voucher": {"tallyvoucher", "vchno", "pdfvouchernumber", "tallyvoucherfound"},
        "match_status": {"matchstatus", "postingstatus", "status"},
        "verified_status": {"verifiedstatus", "verified", "approvalstatus"},
        "final_row": {"finalsheetrow", "incomesheetrow", "excelrow", "finalexcelrow"},
    }


def income_aliases() -> dict[str, set[str]]:
    return {
        "student_code": {"studentcode", "code"},
        "student_name": {"studentname", "name", "namevendor"},
        "course": {"course", "coursename", "program"},
        "amount": {
            "amount",
            "feesreceivedwithgst",
            "grossactualfees",
            "feescollectedwithgst",
            "fp.trainingfeeswithgst",
            "fpatrainingfeeswithgst",
        },
    }


def cell_value(ws, row: int, columns: dict[str, int], key: str) -> Any:
    col = columns.get(key)
    return ws.cell(row, col).value if col else None


def read_recon_records(ws) -> tuple[int, dict[str, int], list[ReconRecord]]:
    header_row, columns = find_header_row(ws, recon_aliases())
    missing = [key for key in ("tally_date", "tally_voucher") if key not in columns]
    if missing:
        raise ValueError("Verified reconciliation sheet must include Tally Date and Tally Voucher columns.")

    records: list[ReconRecord] = []
    for row in range(header_row + 1, ws.max_row + 1):
        tally_date = cell_value(ws, row, columns, "tally_date")
        tally_voucher = cell_value(ws, row, columns, "tally_voucher")
        code = normalize_code(cell_value(ws, row, columns, "student_code"))
        name = str(cell_value(ws, row, columns, "student_name") or "").strip()
        amount = parse_amount(cell_value(ws, row, columns, "amount"))
        if not any([tally_date, tally_voucher, code, name, amount]):
            continue
        final_row_raw = cell_value(ws, row, columns, "final_row")
        try:
            final_row = int(final_row_raw) if final_row_raw not in (None, "") else None
        except (TypeError, ValueError):
            final_row = None
        records.append(
            ReconRecord(
                row_number=row,
                student_code=code,
                student_name=name,
                course=str(cell_value(ws, row, columns, "course") or "").strip(),
                amount=amount,
                tally_date=tally_date,
                tally_voucher=tally_voucher,
                match_status=str(cell_value(ws, row, columns, "match_status") or "").strip(),
                verified_status=str(cell_value(ws, row, columns, "verified_status") or "").strip(),
                final_row=final_row,
            )
        )
    return header_row, columns, records


def read_income_records(ws) -> tuple[int, dict[str, int], list[IncomeRecord]]:
    header_row, columns = find_header_row(ws, income_aliases())
    records: list[IncomeRecord] = []
    for row in range(header_row + 1, ws.max_row + 1):
        code = normalize_code(cell_value(ws, row, columns, "student_code"))
        name = str(cell_value(ws, row, columns, "student_name") or "").strip()
        course = str(cell_value(ws, row, columns, "course") or "").strip()
        amount = parse_amount(cell_value(ws, row, columns, "amount"))
        if code or name or amount:
            records.append(IncomeRecord(row, code, name, course, amount))
    return header_row, columns, records


def validate_candidate(record: ReconRecord, candidate: IncomeRecord, options: PostingOptions) -> list[str]:
    issues: list[str] = []
    if record.student_code and candidate.student_code and record.student_code != candidate.student_code:
        issues.append("Student code mismatch")
    if record.student_name and candidate.student_name and normalize_text(record.student_name) != normalize_text(candidate.student_name):
        issues.append("Student name mismatch")
    if record.course and candidate.course and normalize_text(record.course) != normalize_text(candidate.course):
        issues.append("Course mismatch")
    if options.require_amount_match and record.amount is not None and candidate.amount is not None:
        if not amounts_match(record.amount, candidate.amount):
            issues.append("Amount mismatch")
    return issues


def find_candidates(record: ReconRecord, income_records: list[IncomeRecord], options: PostingOptions) -> tuple[list[IncomeRecord], str]:
    if record.final_row:
        direct = [item for item in income_records if item.row_number == record.final_row]
        return direct, "Final row number"

    code_amount = [
        item for item in income_records if record.student_code and item.student_code == record.student_code and amounts_match(record.amount, item.amount)
    ]
    if code_amount:
        return code_amount, "Priority 1: Student Code + Amount"

    code_name_course = [
        item
        for item in income_records
        if record.student_code
        and item.student_code == record.student_code
        and normalize_text(record.student_name)
        and normalize_text(record.student_name) == normalize_text(item.student_name)
        and (not record.course or not item.course or normalize_text(record.course) == normalize_text(item.course))
    ]
    if code_name_course:
        return code_name_course, "Priority 2: Student Code + Name + Course"

    name_course_amount = [
        item
        for item in income_records
        if normalize_text(record.student_name)
        and normalize_text(record.student_name) == normalize_text(item.student_name)
        and (not record.course or not item.course or normalize_text(record.course) == normalize_text(item.course))
        and amounts_match(record.amount, item.amount)
    ]
    if name_course_amount:
        return name_course_amount, "Priority 3: Name + Course + Amount"

    return [], "No candidate"


def append_report_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 70)


def save_report(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    append_report_rows(ws, headers, rows)
    wb.save(path)


def add_review_marker(ws, row: int, reason: str) -> None:
    for col in (TALLY_DATE_COL, TALLY_VOUCHER_COL):
        cell = ws.cell(row, col)
        cell.fill = WARNING_FILL
        existing = cell.comment.text if cell.comment else ""
        comment_text = f"{existing}\nReview required: {reason}".strip()
        cell.comment = Comment(comment_text[:32000], "FPA Tool")


def prepare_income_sheet_posting(
    recon_path: Path,
    income_path: Path,
    output_dir: Path,
    options: PostingOptions,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    recon_wb = load_workbook(recon_path, data_only=False, read_only=True)
    recon_ws = recon_wb.active
    _, _, recon_records = read_recon_records(recon_ws)

    income_wb = load_workbook(income_path)
    if options.target_sheet not in income_wb.sheetnames:
        raise ValueError(f"Target sheet '{options.target_sheet}' was not found. Available sheets: {', '.join(income_wb.sheetnames)}")
    income_ws = income_wb[options.target_sheet]
    _, _, income_records = read_income_records(income_ws)

    posted_rows: list[list[Any]] = []
    review_rows: list[list[Any]] = []
    summary = {
        "recon_rows_scanned": len(recon_records),
        "income_rows_scanned": len(income_records),
        "rows_posted": 0,
        "rows_skipped": 0,
        "rows_needing_review": 0,
        "existing_values": 0,
        "amount_mismatches": 0,
        "multiple_matches": 0,
        "missing_tally_fields": 0,
    }
    used_income_rows: set[int] = set()

    for record in recon_records:
        fields_checked: list[str] = []
        issue_type = ""
        reason = ""
        target_row: int | None = None
        candidates: list[IncomeRecord] = []

        if not record.tally_date or not record.tally_voucher:
            issue_type = "Missing Tally Date/Voucher"
            reason = "Tally Date or Tally Voucher is blank"
            summary["missing_tally_fields"] += 1
        elif record.match_status and not is_approved_status(record.match_status):
            issue_type = "Match Status Not Approved"
            reason = f"Match Status is '{record.match_status}'"
        elif options.require_verified_status and (not record.verified_status or not is_approved_status(record.verified_status)):
            issue_type = "Verified Status Not Approved"
            reason = f"Verified Status is '{record.verified_status or 'blank'}'"
        else:
            candidates, match_method = find_candidates(record, income_records, options)
            fields_checked.append(match_method)
            if not candidates:
                issue_type = "Target Row Not Found"
                reason = "No unique target row found in income sheet"
            elif len(candidates) > 1:
                issue_type = "Multiple Matches"
                reason = f"{len(candidates)} possible target rows found"
                summary["multiple_matches"] += 1
            else:
                candidate = candidates[0]
                target_row = candidate.row_number
                validation_issues = validate_candidate(record, candidate, options)
                if validation_issues:
                    issue_type = validation_issues[0]
                    reason = "; ".join(validation_issues)
                    if "Amount mismatch" in validation_issues:
                        summary["amount_mismatches"] += 1
                elif target_row in used_income_rows:
                    issue_type = "Duplicate Posting Target"
                    reason = "Another reconciliation row already targeted this income sheet row"
                else:
                    existing_date = income_ws.cell(target_row, TALLY_DATE_COL).value
                    existing_voucher = income_ws.cell(target_row, TALLY_VOUCHER_COL).value
                    if (existing_date or existing_voucher) and not options.allow_overwrite:
                        issue_type = "Existing Tally Values"
                        reason = "Income sheet already has Tally Date or Tally Voucher"
                        summary["existing_values"] += 1
                    else:
                        income_ws.cell(target_row, TALLY_DATE_COL).value = record.tally_date
                        income_ws.cell(target_row, TALLY_VOUCHER_COL).value = record.tally_voucher
                        used_income_rows.add(target_row)
                        summary["rows_posted"] += 1
                        posted_rows.append(
                            [
                                record.row_number,
                                target_row,
                                record.student_code,
                                record.student_name,
                                record.course,
                                record.amount,
                                record.tally_date,
                                record.tally_voucher,
                                "Posted",
                                "Safe posting",
                                ", ".join(fields_checked),
                            ]
                        )
                        continue

        summary["rows_needing_review"] += 1
        summary["rows_skipped"] += 1
        candidate_rows = ""
        candidate_rows = ", ".join(str(item.row_number) for item in candidates)
        review_rows.append(
            [
                record.row_number,
                candidate_rows or target_row or "",
                record.student_code,
                record.student_name,
                record.course,
                record.amount,
                record.tally_date,
                record.tally_voucher,
                issue_type,
                reason,
                "Review manually before posting",
            ]
        )
        if options.highlight_review_rows and target_row:
            add_review_marker(income_ws, target_row, reason)

    updated_path = output_dir / f"{income_path.stem}_posted.xlsx"
    posting_report_path = output_dir / "posting_report.xlsx"
    review_report_path = output_dir / "review_discrepancy_report.xlsx"
    summary_path = output_dir / "posting_summary_report.xlsx"

    income_wb.save(updated_path)
    save_report(
        posting_report_path,
        [
            "Reconciliation Row",
            "Income Sheet Row",
            "Student Code",
            "Student Name",
            "Course",
            "Amount",
            "Tally Date",
            "Tally Voucher",
            "Posting Status",
            "Reason",
            "Fields Checked",
        ],
        posted_rows,
    )
    save_report(
        review_report_path,
        [
            "Reconciliation Row",
            "Candidate Income Sheet Row(s)",
            "Student Code",
            "Student Name",
            "Course",
            "Amount",
            "Tally Date",
            "Tally Voucher",
            "Issue Type",
            "Details",
            "Recommended Action",
        ],
        review_rows,
    )
    save_report(
        summary_path,
        ["Metric", "Value"],
        [
            ["Reconciliation rows scanned", summary["recon_rows_scanned"]],
            ["Income sheet rows scanned", summary["income_rows_scanned"]],
            ["Rows posted", summary["rows_posted"]],
            ["Rows skipped", summary["rows_skipped"]],
            ["Rows needing review", summary["rows_needing_review"]],
            ["Rows with existing Tally Date / Tally Voucher", summary["existing_values"]],
            ["Amount mismatches", summary["amount_mismatches"]],
            ["Multiple match cases", summary["multiple_matches"]],
            ["Missing Tally Date / Tally Voucher cases", summary["missing_tally_fields"]],
            ["Target sheet", options.target_sheet],
            ["Allow overwrite", options.allow_overwrite],
            ["Require verified status", options.require_verified_status],
            ["Require amount match", options.require_amount_match],
        ],
    )

    return {
        **summary,
        "updated_path": updated_path,
        "posting_report_path": posting_report_path,
        "review_report_path": review_report_path,
        "summary_path": summary_path,
        "sheetnames": income_wb.sheetnames,
    }
