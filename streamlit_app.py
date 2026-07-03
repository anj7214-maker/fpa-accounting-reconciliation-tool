from __future__ import annotations

import shutil
from io import BytesIO
from datetime import datetime
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from fpa_monthly_reconcile import export_pdf_tally_details, reconcile
from income_sheet_posting import PostingOptions, prepare_income_sheet_posting


BASE_DIR = Path(__file__).resolve().parent
RUNS_DIR = BASE_DIR / "outputs" / "app_runs"
LOGO_PATH = BASE_DIR / "assets" / "fpa_logo.png"


def safe_name(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "._- " else "_" for ch in name).strip()
    return cleaned or "uploaded_file"


def save_upload(uploaded_file, folder: Path) -> Path:
    path = folder / safe_name(uploaded_file.name)
    path.write_bytes(uploaded_file.getbuffer())
    return path


def save_uploads(uploaded_files, folder: Path) -> list[Path]:
    return [save_upload(uploaded_file, folder) for uploaded_file in uploaded_files]


def read_summary(path: Path) -> dict[str, object]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    return {str(ws.cell(row, 1).value): ws.cell(row, 2).value for row in range(2, ws.max_row + 1)}


def download_button(label: str, path: Path, mime: str) -> None:
    st.download_button(
        label,
        data=path.read_bytes(),
        file_name=path.name,
        mime=mime,
        use_container_width=True,
    )


def uploaded_sheetnames(uploaded_file) -> list[str]:
    if not uploaded_file:
        return []
    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    return workbook.sheetnames


st.set_page_config(page_title="FPA Monthly Reconciliation", page_icon="FPA", layout="wide")

header_logo, header_text = st.columns([1, 6], vertical_alignment="center")
with header_logo:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), width=110)
with header_text:
    st.title("FPA Monthly Reconciliation")
    st.caption("Fill only Tally Date and Tally Voucher, while explaining every Tally PDF voucher.")

extract_tab, reconcile_tab, post_tab = st.tabs(["PDF Extract", "Reconciliation", "Post to Income Sheet"])

with extract_tab:
    st.subheader("Extract Tally Date and Tally Voucher")
    st.caption("Upload one or more Tally PDFs and download one combined Excel file. The first columns are Tally Date and Tally Voucher.")

    extract_pdf_uploads = st.file_uploader(
        "Tally Ledger PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        key="extract_pdf",
    )
    extract_run = st.button("Extract Date and Voucher", type="primary", use_container_width=True)

    if extract_run:
        if not extract_pdf_uploads:
            st.error("Please upload at least one Tally PDF.")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            run_dir = RUNS_DIR / "pdf_extract" / timestamp
            input_dir = run_dir / "inputs"
            output_dir = run_dir / "outputs"
            input_dir.mkdir(parents=True, exist_ok=True)

            pdf_paths = save_uploads(extract_pdf_uploads, input_dir)

            try:
                with st.spinner("Extracting Date and Vch No. from the uploaded PDFs..."):
                    result = export_pdf_tally_details(pdf_paths, output_dir)

                st.success("PDF extraction complete.")
                metric_cols = st.columns(2)
                metric_cols[0].metric("PDF Files", result["pdf_files"])
                metric_cols[1].metric("PDF Transactions Extracted", result["transactions"])
                download_button(
                    "Download Tally Date / Voucher Excel",
                    result["output_path"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as exc:
                st.error("The PDF extraction could not be completed.")
                st.exception(exc)

with reconcile_tab:
    with st.sidebar:
        st.header("Matching")
        match_label = st.radio(
            "Match rule",
            ["Safe mode: code + amount, then unique code", "Review mode: also allow name + amount"],
            index=0,
            help="Safe mode is recommended for normal monthly accounting runs.",
        )
        match_mode = "safe" if match_label.startswith("Safe mode") else "review"

        st.header("Safety")
        st.write("Only these workbook columns are updated:")
        st.code("Tally Date\nTally Voucher")

    left, right = st.columns(2)
    with left:
        pdf_uploads = st.file_uploader(
            "Tally Ledger PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            key="reconcile_pdf",
        )
    with right:
        excel_upload = st.file_uploader("Master Student Accounting Excel", type=["xlsx"])

    run = st.button("Run Reconciliation", type="primary", use_container_width=True)

    if run:
        if not pdf_uploads or not excel_upload:
            st.error("Please upload at least one Tally PDF and the Excel workbook.")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            run_dir = RUNS_DIR / "reconciliation" / timestamp
            input_dir = run_dir / "inputs"
            output_dir = run_dir / "outputs"
            input_dir.mkdir(parents=True, exist_ok=True)

            pdf_paths = save_uploads(pdf_uploads, input_dir)
            excel_path = save_upload(excel_upload, input_dir)

            try:
                with st.spinner("Reconciling PDF vouchers with Excel rows..."):
                    result = reconcile(pdf_paths, excel_path, output_dir, match_mode=match_mode)

                st.success("Reconciliation complete.")

                metric_cols = st.columns(5)
                metric_cols[0].metric("PDF Files", result["pdf_files"])
                metric_cols[1].metric("PDF Matched", result["pdf_matched"])
                metric_cols[2].metric("PDF Unmatched", result["pdf_unmatched"])
                metric_cols[3].metric("PDF Manual Review", result["pdf_manual_review"])
                metric_cols[4].metric("Excel Rows Updated", result["matched"])

                excel_cols = st.columns(4)
                excel_cols[0].metric("PDF Vouchers", result["pdf_transactions"])
                excel_cols[1].metric("Excel Rows Scanned", result["excel_rows"])
                excel_cols[2].metric("Excel Rows Unmatched", result["unmatched"])
                excel_cols[3].metric("Duplicate Review Rows", result["duplicates"])

                summary = read_summary(result["summary_path"])
                with st.expander("Summary details", expanded=False):
                    st.table(summary)

                zip_base = run_dir / "fpa_reconciliation_outputs"
                zip_path = Path(shutil.make_archive(str(zip_base), "zip", output_dir))

                st.subheader("Downloads")
                downloads = st.columns(3)
                with downloads[0]:
                    download_button(
                        "Updated workbook",
                        result["updated_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    download_button(
                        "Matched report",
                        result["matched_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with downloads[1]:
                    download_button(
                        "Excel unmatched report",
                        result["excel_unmatched_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    download_button(
                        "PDF unmatched report",
                        result["pdf_unmatched_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with downloads[2]:
                    download_button(
                        "Duplicate/manual review report",
                        result["duplicate_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    download_button(
                        "Summary report",
                        result["summary_path"],
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    download_button("All outputs as ZIP", zip_path, "application/zip")

            except Exception as exc:
                st.error("The reconciliation could not be completed.")
                st.exception(exc)

with post_tab:
    st.subheader("Post to Income Sheet")
    st.caption("Copy only Tally Date and Tally Voucher from a verified reconciliation sheet into the final Income Sheet.")

    post_left, post_right = st.columns(2)
    with post_left:
        recon_post_upload = st.file_uploader(
            "Verified Reconciliation Sheet",
            type=["xlsx"],
            key="post_recon_workbook",
        )
    with post_right:
        income_upload = st.file_uploader(
            "Final Income Sheet Workbook",
            type=["xlsx"],
            key="post_income_workbook",
        )

    sheetnames = []
    if income_upload:
        try:
            sheetnames = uploaded_sheetnames(income_upload)
        except Exception as exc:
            st.warning(f"Could not read income workbook sheets yet: {exc}")

    default_sheet_index = sheetnames.index("IS25-26") if "IS25-26" in sheetnames else 0
    target_sheet = st.selectbox(
        "Target income sheet",
        sheetnames or ["IS25-26"],
        index=default_sheet_index if sheetnames else 0,
        help="The tool prefers IS25-26. Choose another sheet only if needed.",
    )

    option_cols = st.columns(4)
    allow_overwrite = option_cols[0].checkbox("Allow overwrite", value=False)
    highlight_review = option_cols[1].checkbox("Highlight review rows", value=True)
    require_verified = option_cols[2].checkbox("Require verified status", value=True)
    require_amount = option_cols[3].checkbox("Require amount match", value=True)

    st.info("Only columns AK and AL are posted for safe rows. Review rows are not posted automatically.")

    post_run = st.button("Preview Posting", type="primary", use_container_width=True)

    if post_run:
        if not recon_post_upload or not income_upload:
            st.error("Please upload both the verified reconciliation sheet and the final income sheet workbook.")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            run_dir = RUNS_DIR / "income_posting" / timestamp
            input_dir = run_dir / "inputs"
            output_dir = run_dir / "outputs"
            input_dir.mkdir(parents=True, exist_ok=True)

            recon_path = save_upload(recon_post_upload, input_dir)
            income_path = save_upload(income_upload, input_dir)
            options = PostingOptions(
                target_sheet=target_sheet,
                allow_overwrite=allow_overwrite,
                highlight_review_rows=highlight_review,
                require_verified_status=require_verified,
                require_amount_match=require_amount,
            )

            try:
                with st.spinner("Preparing controlled income-sheet posting..."):
                    result = prepare_income_sheet_posting(recon_path, income_path, output_dir, options)

                zip_base = run_dir / "income_posting_outputs"
                result["zip_path"] = Path(shutil.make_archive(str(zip_base), "zip", output_dir))
                st.session_state["income_posting_result"] = result
                st.session_state["income_posting_downloads_ready"] = False
            except Exception as exc:
                st.error("The income-sheet posting could not be completed.")
                st.exception(exc)

    posting_result = st.session_state.get("income_posting_result")
    if posting_result:
        st.success("Posting preview is ready. Review the metrics before downloading the updated income sheet.")
        metrics = st.columns(5)
        metrics[0].metric("Rows Ready to Post", posting_result["rows_posted"])
        metrics[1].metric("Rows Needing Review", posting_result["rows_needing_review"])
        metrics[2].metric("Rows Skipped", posting_result["rows_skipped"])
        metrics[3].metric("Existing Values", posting_result["existing_values"])
        metrics[4].metric("Amount Mismatches", posting_result["amount_mismatches"])

        extra_metrics = st.columns(3)
        extra_metrics[0].metric("Recon Rows Scanned", posting_result["recon_rows_scanned"])
        extra_metrics[1].metric("Income Rows Scanned", posting_result["income_rows_scanned"])
        extra_metrics[2].metric("Multiple Matches", posting_result["multiple_matches"])

        if st.button("Generate Updated Income Sheet Downloads", use_container_width=True):
            st.session_state["income_posting_downloads_ready"] = True

        if st.session_state.get("income_posting_downloads_ready"):
            st.subheader("Downloads")
            downloads = st.columns(2)
            with downloads[0]:
                download_button(
                    "Updated Income Sheet",
                    posting_result["updated_path"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                download_button(
                    "Posting Report",
                    posting_result["posting_report_path"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with downloads[1]:
                download_button(
                    "Review / Discrepancy Report",
                    posting_result["review_report_path"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                download_button(
                    "Summary Report",
                    posting_result["summary_path"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                download_button("All posting outputs as ZIP", posting_result["zip_path"], "application/zip")
